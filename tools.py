# -*- coding: utf-8 -*-

import datetime
import calendar
import time
import os
import shutil
import urllib3
import requests
import platform
import pandas as pd
import numpy as np
import hashlib
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


from strings import *
from parameters import *


def http_get(cookie, url):
    urllib3.disable_warnings()
    headers = {
        STRING_COOKIE: cookie
    }
    try:
        response = requests.get(url, headers=headers, verify=False)
        response.encoding = STRING_UTF_8
        if response.status_code == 200:
            return clear(response.text, TERRIBLE_XML)
        return None
    except requests.exceptions.RequestException:
        return None


def clear(string, terrible=None, trans=None):
    if terrible is None:
        terrible = TERRIBLE_DEFAULT
    if trans is None:
        trans = TRANS_DEFAULT
    for item in terrible:
        string = string.replace(item, STRING_EMPTY)
    for item in trans:
        string = string.replace(item[0], item[1])
    return string


def clear_name(string):
    string = string.split(STRING_LEFT_SQUARE_BRACKET)[0]
    string = clear(string, TERRIBLE_NAMES)
    return string


def no_bad(string, bad):
    for bad_word in bad:
        if bad_word in string:
            return False
    return True


def last_legal_period(count, time_string):
    # print("original stamp:", string_to_stamp(time_string, STRING_FORMAT_FULL_SECOND))
    last_friday = datetime.date.fromtimestamp(string_to_stamp(time_string, STRING_FORMAT_FULL_SECOND))
    while last_friday.weekday() != calendar.FRIDAY:
        last_friday -= datetime.timedelta(days=1)
    last_second_friday = last_friday - datetime.timedelta(days=7) * count
    last_friday_stamp = time.mktime(last_friday.timetuple())
    last_second_friday_stamp = time.mktime(last_second_friday.timetuple())
    return last_second_friday_stamp, last_friday_stamp


def string_to_stamp(string, string_format=STRING_FORMAT_FULL_MINUTE):
    # if platform.system().lower() == "linux":
    #     return time.mktime(time.strptime(utc_to_local(string, string_format), string_format))
    return time.mktime(time.strptime(string, string_format))


def stamp_to_string(stamp, string_format=STRING_FORMAT_FULL_MINUTE):
    # if platform.system().lower() == "linux":
    #     return utc_to_local(time.strftime(string_format, time.localtime(stamp)), string_format)
    return time.strftime(string_format, time.localtime(stamp))


def print_df(df):
    pd.set_option(STRING_DISPLAY_MAX_ROWS, 500)
    pd.set_option(STRING_DISPLAY_MAX_COLUMNS, 500)
    pd.set_option(STRING_DISPLAY_WIDTH, 1000)
    print(df)
    

def select_first_match(df, key_col_1, key_1, key_col_2, key_2, res_col):
    df = df[(df[key_col_1] == key_1) & (df[key_col_2] == key_2)]
    res = df.copy(deep=True)
    if len(res) == 0:
        return np.nan
    return res.iloc[0][res_col]


def sum_all_match(df, key_col, key, res_col):
    df = df[df[key_col] == key]
    res = df.copy(deep=True)
    if len(res) == 0:
        return 0
    res_sum = 0.0
    for i in range(0, len(res)):
        res_sum += res.iloc[i][res_col]
    return res_sum


def not_all_nan(df_col):
    for item in df_col:
        if not np.isnan(item):
            return True
    return False


def not_all_like(df_col, like):
    for item in df_col:
        if item != like:
            return True
    return False


def mean_satis(df_col):
    if not_all_like(df_col, STRING_CN_NONE):
        satis_sum = 0
        satis_count = 0
        for item in df_col:
            if item != STRING_CN_NONE:
                satis_sum += int(item)
                satis_count += 1
        return satis_sum / float(satis_count)
    else:
        return STRING_CN_NONE
    

def create_path(execute_path, legal_start, legal_end):
    start_string = stamp_to_string(legal_start, STRING_FORMAT_NO_HYPHEN)
    end_string = stamp_to_string(legal_end, STRING_FORMAT_NO_HYPHEN)
    dir_path = STRING_FORMAT_7.format(
        execute_path, os.sep, STRING_SAVES, os.sep, start_string, STRING_HYPHEN, end_string)
    version_string = stamp_to_string(time.time(), STRING_FORMAT_FULL_SECOND_NO_HYPHEN)
    filename = STRING_FORMAT_FILENAME.format(start_string, end_string, version_string)
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    full_path = STRING_FORMAT_3.format(dir_path, os.sep, filename)
    return full_path, filename


def clean_dir(path):
    files = os.listdir(path)
    for file in files:
        full_path = STRING_FORMAT_3.format(path, os.sep, file)
        if os.path.isfile(full_path):
            os.remove(full_path)


def copy_file(from_path, to_path):
    if os.path.exists(from_path) and os.path.isfile(from_path):
        shutil.copyfile(from_path, to_path)
        
        
def now_time_string():
    # if platform.system().lower() == "linux":
    #     return utc_to_local(time.strftime(STRING_FORMAT_FULL_SECOND, time.gmtime()))
    return stamp_to_string(time.time(), STRING_FORMAT_FULL_SECOND)


def get_week_day(string, string_format=STRING_FORMAT_FULL_SECOND):
    return datetime.datetime.fromtimestamp(time.mktime(time.strptime(string, string_format))).isoweekday()


class LogFBR:
    def __init__(self):
        self.logs = []
        self.flag = False
        self.log = STRING_EMPTY
    
    def print(self, string, log_end=None):
        if self.flag:
            full_log = string
        else:
            full_log = STRING_FORMAT_3.format(now_time_string(), STRING_SPACE, string)
        if log_end is not None:
            print(full_log, end=STRING_EMPTY)
            self.logs.append(full_log)
            self.flag = True
        else:
            print(full_log)
            full_log += STRING_PLAIN_ENTER
            self.logs.append(full_log)
            self.flag = False
        self.log += full_log
        

def sort_time_department(src_df):
    df = src_df.copy(deep=True)
    df["负参与流程数"] = - 1 * df["参与流程数"]
    df["部门字数"] = [len(x) for x in df["所在部门"]]
    df = df.sort_values(by=["负参与流程数", "部门字数", "所在部门"], ascending=True)
    del df["负参与流程数"]
    del df["部门字数"]
    return df.reset_index(drop=True)


def sort_time_person(src_df):
    df = src_df.copy(deep=True)
    df["负平均处理时长（小时）"] = - 1 * df["平均处理时长（小时）"]
    df = df.sort_values(by=["负平均处理时长（小时）", "操作者"], ascending=True)
    del df["负平均处理时长（小时）"]
    return df.reset_index(drop=True)


def sort_summary(src_df):
    df = src_df.copy(deep=True)
    df["负参与流程数"] = - 1 * df["解决需求数"]
    df["部门字数"] = [len(x) for x in df["所在部门"]]
    df = df.sort_values(by=["负参与流程数", "部门字数", "所在部门"], ascending=True)
    del df["负参与流程数"]
    del df["部门字数"]
    return df.reset_index(drop=True)


def sha256(filename):
    with open(filename, "rb") as f:
        data = f.read()
    file_md5 = hashlib.sha256(data).hexdigest()
    return file_md5


def unite_html(content_list):
    content_html = STRING_MAIL_TEXT_HEAD
    for item in content_list:
        if item[0] == "log":
            new_content = STRING_MAIL_TEXT_PART_LOG.format(
                item[1],
                item[2].replace(STRING_MAIL_TEXT_ALIGN_RIGHT, "").replace("NaN", "")
            )
            content_html += new_content
        elif item[0] == "highlight":
            new_content = STRING_MAIL_TEXT_PART_NORMAL.format(
                item[1],
                item[2].replace(STRING_MAIL_TEXT_ALIGN_RIGHT, "").replace("NaN", "")
            )
            content_html += new_content.replace("class=\"dataframe\"", "class=\"dataframe\" id=\"table_highlight\"")
        elif item[0] == "none":
            new_content = STRING_MAIL_TEXT_PART_NONE.format(
                item[1],
                STRING_CN_NO_CONTENT_WARNING
            )
            content_html += new_content
        else:
            new_content = STRING_MAIL_TEXT_PART_NORMAL.format(
                item[1],
                item[2].replace(STRING_MAIL_TEXT_ALIGN_RIGHT, "")
            )
            content_html += new_content.replace("class=\"dataframe\"", "class=\"dataframe\" id=\"table_normal\"")
    content_html += STRING_MAIL_TEXT_TAIL
    return content_html


def set_excel_style(filename):
    book = openpyxl.load_workbook(filename)
    # book.active = 0
    # 初始化字体样式
    font_bold = Font(name='等线',
                     size=11,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='FF000000',
                     outline='None')

    font_not_bold = Font(name='等线',
                         size=11,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000',
                         outline='None')

    font_link = Font(name='等线',
                     size=11,
                     italic=False,
                     vertAlign=None,
                     underline='single',
                     strike=False,
                     color='1810d2',
                     outline='None')

    border_none = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    border_full = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'),
                         left=Side(border_style='thin'), right=Side(border_style='thin'))
    border_double = Border(top=Side(border_style='thin'), bottom=Side(border_style='double'), left=Side(), right=Side())
    for one_sheet_name in book.sheetnames:
        sheet: Worksheet = book[one_sheet_name]
        # if sheet.title in not_empty_business_list:
        #     tmp_max_column = sheet.max_column
        #     sheet.cell(1, tmp_max_column + 1).value = STRING_COL_REGISTRATION_COUNT
        #     sheet.cell(1, tmp_max_column + 2).value = STRING_COL_FILE_COUNT
        #     sheet.cell(1, tmp_max_column + 3).value = STRING_COL_PAGE_COUNT
        max_row = sheet.max_row
        max_column = sheet.max_column


        # 设置整体的字体
        for row in sheet.rows:
            for cell in row:
                cell.font = font_not_bold
                cell.border = border_full

        # 加粗
        for i in range(1, sheet.max_column + 1):
            sheet.cell(1, i).font = font_bold

        # 居中
        for i in range(max_row):
            for j in range(max_column):
                if i == 0:
                    sheet.cell(i + 1, j + 1).fill = PatternFill(fill_type='solid', fgColor="B4C6E7")
                sheet.cell(i + 1, j + 1).alignment = Alignment(horizontal='center', vertical='center')


        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        # 遍历全部列
        for i in range(1, max_column + 1):
            # 遍历每一列的全部行
            for j in range(1, max_row + 1):
                column = 0
                # 获取j行i列的值
                sheet_value = sheet.cell(j, i).value
                if "!A1" in str(sheet_value):
                    sheet_value = str(sheet_value).split("\"")[3]
                # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                sheet_value_list = [k for k in str(sheet_value)]
                # print(sheet_value_list, end="")
                # 遍历当前单元格的字符列表
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    # if v.isdigit() or v.isalpha():
                    if ord(v) < 256:
                        column += 1.3
                    else:
                        column += 2.2
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                # print(column)
                try:
                    if not max_column_dict.get(i) or column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    print("Error:", e)
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[get_column_letter(key)].width = value

        book.save(filename)


# def utc_to_local(utc_time_str, utc_format=STRING_FORMAT_FULL_SECOND, timezone="Asia/Shanghai"):
#     local_tz = pytz.timezone(timezone)
#     local_format = utc_format
#     utc_dt = datetime.datetime.strptime(utc_time_str, utc_format)
#     local_dt = utc_dt.replace(tzinfo=pytz.utc).astimezone(local_tz)
#     time_str = local_dt.strftime(local_format)
#     return time_str


if __name__ == "__main__":
    print(sha256(r"D:\Workspace\AI_test\pics\other_validation\other_9.jpg"))
    

